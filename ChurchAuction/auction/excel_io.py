"""Excel import (catalog in) and export (reports out).

Excel is used only at the edges of an event, never as the live store. Prices
are written as real numbers with a currency *number format* so the file stays
clean and re-importable (the old code stored strings like "$45.0").
"""

from openpyxl import Workbook, load_workbook

from auction.auction import Auction
from auction.items import Item

CURRENCY_FORMAT = '"$"#,##0.00'


def _header_map(header_row) -> dict[str, int]:
    """Map lower-cased header names to column indices, tolerating spacing."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower().replace(" ", "")
        mapping[key] = idx
    return mapping


def import_items(path: str, sheet_name: str = "Items") -> list[Item]:
    """Read the item catalog from an Excel file.

    Expects columns named ItemId/ItemNumber, Name, and Type (case/space
    insensitive). Any sale columns present are ignored.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    cols = _header_map(header)
    id_col = cols.get("itemid", cols.get("itemnumber", 0))
    name_col = cols.get("name", 1)
    type_col = cols.get("type", cols.get("itemtype", 2))

    items: list[Item] = []
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        raw_id = row[id_col]
        if raw_id is None:
            continue
        item_type = row[type_col] if type_col < len(row) and row[type_col] else "silent"
        items.append(
            Item(
                itemNumber=int(raw_id),
                name=str(row[name_col]) if name_col < len(row) else "",
                itemType=str(item_type).strip().lower(),
            )
        )
    return items


def _bidder_status(bidder) -> str:
    if bidder.fullyPaid:
        return "Paid in full"
    if bidder.amountPaid > 0:
        return "Partial"
    return "Due"


def export_auction(auction: Auction, path: str) -> None:
    """Write Items, Bidders, and Outstanding sheets with numeric prices."""
    wb = Workbook()

    items_sheet = wb.active
    items_sheet.title = "Items"
    bidders_sheet = wb.create_sheet("Bidders")
    outstanding_sheet = wb.create_sheet("Outstanding")

    items_sheet.append(["ItemId", "Name", "Type", "SalePrice", "WinnerId"])
    bidders_sheet.append(
        ["BidderId", "Name", "Items Won", "Item Numbers", "TotalOwed", "AmountPaid", "BalanceDue", "Status"]
    )
    outstanding_sheet.append(
        ["BidderId", "Name", "Items To Deliver", "Item Numbers", "Amount Still Owed"]
    )

    for item in auction.items.values():
        items_sheet.append(
            [item.itemNumber, item.name, item.itemType, item.salePrice, item.winnerId]
        )
        # SalePrice is column D; format the cell we just wrote.
        if item.salePrice is not None:
            items_sheet.cell(row=items_sheet.max_row, column=4).number_format = CURRENCY_FORMAT

    for bidder in auction.bidders.values():
        won = ", ".join(
            auction.items[i].name for i in bidder.itemsWon if i in auction.items
        )
        numbers = ", ".join(str(i) for i in bidder.itemsWon)
        bidders_sheet.append(
            [
                bidder.bidderId,
                bidder.name,
                won,
                numbers,
                bidder.totalOwed,
                bidder.amountPaid,
                bidder.balanceDue,
                _bidder_status(bidder),
            ]
        )
        row = bidders_sheet.max_row
        for col in (5, 6, 7):  # TotalOwed, AmountPaid, BalanceDue
            bidders_sheet.cell(row=row, column=col).number_format = CURRENCY_FORMAT

    # Post-auction to-do list: anyone who still owes money or has items to hand over.
    for bidder in auction.bidders.values():
        outstanding = bidder.outstandingItems
        if bidder.balanceDue <= 0 and not outstanding:
            continue
        names = ", ".join(
            auction.items[i].name for i in outstanding if i in auction.items
        )
        numbers = ", ".join(str(i) for i in outstanding)
        outstanding_sheet.append(
            [bidder.bidderId, bidder.name, names, numbers, bidder.balanceDue]
        )
        outstanding_sheet.cell(row=outstanding_sheet.max_row, column=5).number_format = CURRENCY_FORMAT

    wb.save(path)
